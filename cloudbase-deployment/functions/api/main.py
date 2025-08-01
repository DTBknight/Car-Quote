#!/usr/bin/env python3
"""
腾讯云CloudBase云函数 - Flask应用入口
"""
import sys
import os
import io
import json
import logging
from datetime import datetime
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 创建Flask应用
app = Flask(__name__)
CORS(app, origins=["*"])

# 模板文件路径
TEMPLATE_PATH = 'api/template.xlsx'

@app.route('/health', methods=['GET'])
def health_check():
    """健康检查接口"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'service': 'car-quote-api',
        'platform': 'tencent-cloudbase'
    })

@app.route('/api/generate-contract', methods=['POST'])
def generate_contract():
    """生成合同文件"""
    try:
        # 获取请求数据
        data = request.get_json()
        if not data:
            return jsonify({'error': '无效的请求数据'}), 400
        
        logger.info(f"收到合同生成请求: {data.get('contractType', 'unknown')}")
        
        # 验证必要字段
        required_fields = ['contractType', 'basicInfo', 'goodsData']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'缺少必要字段: {field}'}), 400
        
        # 生成合同文件
        excel_data = generate_excel_contract(data)
        
        # 返回文件
        return send_file(
            io.BytesIO(excel_data),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'contract_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"生成合同失败: {str(e)}")
        return jsonify({'error': f'生成合同失败: {str(e)}'}), 500

def generate_excel_contract(data):
    """生成Excel合同文件"""
    try:
        # 加载模板
        workbook = openpyxl.load_workbook(TEMPLATE_PATH)
        
        # 获取工作表
        sc_sheet = workbook['销售合同']
        pi_sheet = workbook['形式发票']
        
        # 设置基础信息
        set_basic_info(sc_sheet, pi_sheet, data['basicInfo'])
        
        # 设置货物数据
        set_goods_data(sc_sheet, pi_sheet, data['goodsData'])
        
        # 设置其他字段
        if 'otherFields' in data:
            set_other_fields(sc_sheet, pi_sheet, data['otherFields'])
        
        # 保存到内存
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        logger.info("合同文件生成成功")
        return output.getvalue()
        
    except Exception as e:
        logger.error(f"生成Excel文件失败: {e}")
        raise

def set_basic_info(sc_sheet, pi_sheet, basic_info):
    """设置基础信息"""
    try:
        buyer_name = basic_info.get('buyerName', '')
        seller_name = basic_info.get('sellerName', '')
        contract_date = basic_info.get('contractDate', '')
        contract_number = basic_info.get('contractNumber', '')
        
        for sheet in [sc_sheet, pi_sheet]:
            try:
                # 买方名称 - C3-D3合并单元格
                if sheet['C3'].value is not None or buyer_name:
                    sheet['C3'].value = buyer_name
                logger.info(f"设置单元格 C3 成功: {buyer_name}")
            except Exception as e:
                logger.error(f"设置单元格 C3 失败: {e}")
            
            try:
                # 卖方名称 - C4-D4合并单元格
                if sheet['C4'].value is not None or seller_name:
                    sheet['C4'].value = seller_name
                logger.info(f"设置单元格 C4 成功: {seller_name}")
            except Exception as e:
                logger.error(f"设置单元格 C4 失败: {e}")
            
            try:
                # 合同日期 - G3-H3合并单元格
                if sheet['G3'].value is not None or contract_date:
                    sheet['G3'].value = contract_date
                logger.info(f"设置单元格 G3 成功: {contract_date}")
            except Exception as e:
                logger.error(f"设置单元格 G3 失败: {e}")
            
            try:
                # 合同编号 - G4-H4合并单元格
                if sheet['G4'].value is not None or contract_number:
                    sheet['G4'].value = contract_number
                logger.info(f"设置单元格 G4 成功: {contract_number}")
            except Exception as e:
                logger.error(f"设置单元格 G4 失败: {e}")
        
        logger.info("基础信息设置成功")
        
    except Exception as e:
        logger.error(f"设置基础信息失败: {e}")
        raise

def set_goods_data(sc_sheet, pi_sheet, goods_data):
    """设置货物数据"""
    try:
        logger.info(f"收到的货物数据: {goods_data}")
        
        for i, item in enumerate(goods_data, start=1):
            row = 10 + i  # 从第11行开始
            
            for sheet in [sc_sheet, pi_sheet]:
                try:
                    # 序号
                    if sheet[f'B{row}'].value is not None or item.get('id'):
                        sheet[f'B{row}'].value = item.get('id', i)
                except Exception as e:
                    logger.error(f"设置单元格 B{row} 失败: {e}")
                
                try:
                    # 型号
                    if sheet[f'C{row}'].value is not None or item.get('model'):
                        sheet[f'C{row}'].value = item.get('model', '')
                except Exception as e:
                    logger.error(f"设置单元格 C{row} 失败: {e}")
                
                try:
                    # 描述
                    if sheet[f'D{row}'].value is not None or item.get('description'):
                        sheet[f'D{row}'].value = item.get('description', '')
                except Exception as e:
                    logger.error(f"设置单元格 D{row} 失败: {e}")
                
                try:
                    # 颜色
                    if sheet[f'E{row}'].value is not None or item.get('color'):
                        sheet[f'E{row}'].value = item.get('color', '')
                except Exception as e:
                    logger.error(f"设置单元格 E{row} 失败: {e}")
                
                try:
                    # 数量
                    if sheet[f'F{row}'].value is not None or item.get('quantity'):
                        sheet[f'F{row}'].value = item.get('quantity', 0)
                except Exception as e:
                    logger.error(f"设置单元格 F{row} 失败: {e}")
                
                try:
                    # 单价
                    if sheet[f'G{row}'].value is not None or item.get('unitPrice'):
                        sheet[f'G{row}'].value = item.get('unitPrice', 0)
                except Exception as e:
                    logger.error(f"设置单元格 G{row} 失败: {e}")
                
                try:
                    # 总金额
                    if sheet[f'H{row}'].value is not None or item.get('totalAmount'):
                        sheet[f'H{row}'].value = item.get('totalAmount', 0)
                except Exception as e:
                    logger.error(f"设置单元格 H{row} 失败: {e}")
        
        logger.info("货物数据设置成功")
        
    except Exception as e:
        logger.error(f"设置货物数据失败: {e}")
        raise

def set_other_fields(sc_sheet, pi_sheet, other_fields):
    """设置其他字段"""
    try:
        # 运输信息
        f22_value = other_fields.get('f22', '')
        logger.info(f"收到的F22值: {f22_value}")
        
        if f22_value:
            logger.info(f"已添加F22值到映射: {f22_value}")
            for sheet in [sc_sheet, pi_sheet]:
                try:
                    if sheet['F22'].value is not None or f22_value:
                        sheet['F22'].value = f22_value
                    logger.info(f"设置单元格 F22 成功: {f22_value}")
                except Exception as e:
                    logger.error(f"设置单元格 F22 失败: {e}")
        
        # 其他字段映射
        field_mappings = {
            'b23': 'B23',
            'd24': 'D24',
            'd26': 'D26'
        }
        
        for field_key, cell_ref in field_mappings.items():
            if field_key in other_fields:
                value = other_fields[field_key]
                for sheet in [sc_sheet, pi_sheet]:
                    try:
                        if sheet[cell_ref].value is not None or value:
                            sheet[cell_ref].value = value
                        logger.info(f"设置单元格 {cell_ref} 成功: {value}")
                    except Exception as e:
                        logger.error(f"设置单元格 {cell_ref} 失败: {e}")
        
        logger.info("其他字段设置成功")
        
    except Exception as e:
        logger.error(f"设置其他字段失败: {e}")
        raise

# CloudBase云函数入口
def main_handler(event, context):
    """云函数入口点"""
    try:
        # 解析事件
        if 'httpMethod' in event:
            # HTTP触发器
            return handle_http_request(event, context)
        else:
            # 其他触发器
            return {
                'statusCode': 400,
                'body': json.dumps({'error': '不支持的触发器类型'})
            }
    except Exception as e:
        logger.error(f"云函数执行失败: {e}")
        return {
            'statusCode': 500,
            'body': json.dumps({'error': f'服务器内部错误: {str(e)}'})
        }

def handle_http_request(event, context):
    """处理HTTP请求"""
    try:
        # 构建Flask请求环境
        path = event.get('path', '/')
        http_method = event.get('httpMethod', 'GET')
        headers = event.get('headers', {})
        query_string = event.get('queryString', '')
        body = event.get('body', '')
        
        # 创建Flask请求
        with app.test_request_context(
            path=path,
            method=http_method,
            headers=headers,
            query_string=query_string,
            data=body
        ):
            # 执行Flask应用
            response = app.full_dispatch_request()
            
            # 返回响应
            return {
                'statusCode': response.status_code,
                'headers': dict(response.headers),
                'body': response.get_data(as_text=True)
            }
            
    except Exception as e:
        logger.error(f"处理HTTP请求失败: {e}")
        return {
            'statusCode': 500,
            'body': json.dumps({'error': f'处理请求失败: {str(e)}'})
        }

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000))) 