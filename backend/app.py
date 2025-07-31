from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime
import logging

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)  # 允许跨域请求

# 配置
TEMPLATE_PATH = 'template.xlsx'
OUTPUT_DIR = 'outputs'

# 确保输出目录存在
os.makedirs(OUTPUT_DIR, exist_ok=True)

@app.route('/health', methods=['GET'])
def health_check():
    """健康检查接口"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'message': '合同管理后端服务运行正常'
    })

@app.route('/generate-contract', methods=['POST'])
def generate_contract():
    """生成合同Excel文件"""
    try:
        # 获取请求数据
        data = request.get_json()
        logger.info(f"收到合同数据: {data}")
        
        if not data:
            return jsonify({'error': '未收到数据'}), 400
        
        # 加载模板文件
        if not os.path.exists(TEMPLATE_PATH):
            return jsonify({'error': '模板文件不存在'}), 404
        
        # 使用openpyxl加载工作簿
        workbook = load_workbook(TEMPLATE_PATH)
        worksheet = workbook.active
        
        # 填充买方信息
        buyer_name = data.get('buyerName', '')
        buyer_phone = data.get('buyerPhone', '')
        buyer_address = data.get('buyerAddress', '')
        
        # 填充卖方信息
        seller_name = data.get('sellerName', '')
        seller_phone = data.get('sellerPhone', '')
        seller_address = data.get('sellerAddress', '')
        
        # 填充合同信息
        contract_number = data.get('contractNumber', '')
        contract_date_raw = data.get('contractDate', '')
        contract_location = data.get('contractLocation', '')
        
        # 处理日期格式：将 YYYY-MM-DD 转换为 YYYY.MM.DD
        if contract_date_raw:
            try:
                # 解析日期并重新格式化
                from datetime import datetime
                date_obj = datetime.strptime(contract_date_raw, '%Y-%m-%d')
                contract_date = date_obj.strftime('%Y.%m.%d')
            except:
                contract_date = contract_date_raw
        else:
            # 如果没有日期，使用当前日期
            from datetime import datetime
            contract_date = datetime.now().strftime('%Y.%m.%d')
        
        # 填充开户行信息
        bank_info = data.get('bankInfo', '')
        
        # 填充Excel单元格
        # 买方名称 - C3-D3合并单元格
        worksheet['C3'] = buyer_name
        
        # 买方地址 - C4-D4合并单元格
        worksheet['C4'] = buyer_address
        
        # 买方电话 - C5-D5合并单元格
        worksheet['C5'] = buyer_phone
        
        # 卖方名称 - C6-D6合并单元格
        worksheet['C6'] = seller_name
        
        # 卖方地址 - C7-D7合并单元格
        worksheet['C7'] = seller_address
        
        # 卖方电话 - C8-D8合并单元格
        worksheet['C8'] = seller_phone
        
        # 合同编号 - G3单元格
        worksheet['G3'] = contract_number
        
        # 合同日期 - G4单元格
        worksheet['G4'] = contract_date
        
        # 合同地点 - G5单元格
        worksheet['G5'] = contract_location
        
        # 开户行信息 - E7-G8合并单元格
        worksheet['E7'] = bank_info
        
        # 处理新增字段
        export_type = data.get('exportType', '')
        payment_terms = data.get('paymentTerms', '')
        total_amount = data.get('totalAmount', 0)
        amount_in_words = data.get('amountInWords', '')
        f22_value = data.get('f22Value', '')
        port_of_loading = data.get('portOfLoading', '')
        final_destination = data.get('finalDestination', '')
        transport_route = data.get('transportRoute', '')
        mode_of_shipment = data.get('modeOfShipment', '')
        
        # 填充新增字段到Excel
        # F22 - 出口类型 + 港口/国家
        worksheet['F22'] = f22_value
        # D24-G24 - 支付条款
        worksheet['D24'] = payment_terms
        # G21 - 总价
        worksheet['G21'] = total_amount
        # B23-G23 - 大写金额（合并单元格）
        worksheet['B23'] = amount_in_words
        # D21-E21 - 起运港
        worksheet['D21'] = port_of_loading
        # D22-E22 - 目的国
        worksheet['D22'] = final_destination
        # D25-G25 - 运输路线
        worksheet['D25'] = transport_route
        # D26-G26 - 运输方式
        worksheet['D26'] = mode_of_shipment
        
        # 处理货物信息 - 同时处理SC和PI两个sheet
        goods_data = data.get('goodsData', [])
        logger.info(f"收到的货物数据: {goods_data}")
        if goods_data:
            # 获取两个sheet
            sc_sheet = workbook['SC']  # SC sheet
            pi_sheet = workbook['PI']  # PI sheet
            
            # 先处理行的显示/隐藏逻辑
            goods_count = len(goods_data)
            
            # 默认隐藏所有货物行（第11-20行）在两个sheet中
            for row_num in range(11, 21):
                sc_sheet.row_dimensions[row_num].hidden = True
                pi_sheet.row_dimensions[row_num].hidden = True
            
            # 根据实际货物行数显示对应的行在两个sheet中
            for i in range(min(goods_count, 10)):  # 最多支持10行货物
                row_num = 11 + i
                sc_sheet.row_dimensions[row_num].hidden = False
                pi_sheet.row_dimensions[row_num].hidden = False
            
            # 填充货物信息到对应的行在两个sheet中
            for i, goods in enumerate(goods_data):
                if i >= 10:  # 最多支持10行货物
                    break
                    
                current_row = 11 + i
                
                # 在SC sheet中填充货物信息
                sc_sheet[f'B{current_row}'] = goods.get('model', '')  # 型号
                sc_sheet[f'C{current_row}'] = goods.get('description', '')  # 货物名称及规格
                sc_sheet[f'D{current_row}'] = goods.get('color', '')  # 颜色
                sc_sheet[f'E{current_row}'] = goods.get('quantity', 0)  # 数量
                sc_sheet[f'F{current_row}'] = goods.get('unitPrice', 0)  # 单价
                sc_sheet[f'G{current_row}'] = goods.get('totalAmount', 0)  # 金额
                
                # 在PI sheet中填充货物信息
                pi_sheet[f'B{current_row}'] = goods.get('model', '')  # 型号
                pi_sheet[f'C{current_row}'] = goods.get('description', '')  # 货物名称及规格
                pi_sheet[f'D{current_row}'] = goods.get('color', '')  # 颜色
                pi_sheet[f'E{current_row}'] = goods.get('quantity', 0)  # 数量
                pi_sheet[f'F{current_row}'] = goods.get('unitPrice', 0)  # 单价
                pi_sheet[f'G{current_row}'] = goods.get('totalAmount', 0)  # 金额
        
        # 生成输出文件名 - 使用合同编号
        if contract_number:
            # 清理合同编号中的特殊字符，确保文件名安全
            safe_contract_number = "".join(c for c in contract_number if c.isalnum() or c in ('-', '_'))
            output_filename = f'{safe_contract_number}_Contract.xlsx'
        else:
            # 如果没有合同编号，使用时间戳作为备用
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f'contract_{timestamp}.xlsx'
        
        output_path = os.path.join(OUTPUT_DIR, output_filename)
        
        # 保存文件
        workbook.save(output_path)
        logger.info(f"合同文件已生成: {output_path}")
        
        # 返回文件下载
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"生成合同时发生错误: {str(e)}")
        return jsonify({'error': f'生成合同失败: {str(e)}'}), 500

@app.route('/template-info', methods=['GET'])
def get_template_info():
    """获取模板信息"""
    try:
        if not os.path.exists(TEMPLATE_PATH):
            return jsonify({'error': '模板文件不存在'}), 404
        
        workbook = load_workbook(TEMPLATE_PATH)
        worksheet = workbook.active
        
        # 获取模板基本信息
        info = {
            'filename': TEMPLATE_PATH,
            'sheets': workbook.sheetnames,
            'max_row': worksheet.max_row,
            'max_column': worksheet.max_column,
            'template_cells': {
                'buyer_name': 'C3',
                'buyer_address': 'C4', 
                'buyer_phone': 'C5',
                'seller_name': 'C6',
                'seller_address': 'C7',
                'seller_phone': 'C8',
                'contract_number': 'G3',
                'contract_date': 'G4',
                'contract_location': 'G5',
                'bank_info': 'E7'
            }
        }
        
        return jsonify(info)
        
    except Exception as e:
        logger.error(f"获取模板信息时发生错误: {str(e)}")
        return jsonify({'error': f'获取模板信息失败: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001) 