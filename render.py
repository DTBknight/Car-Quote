from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime
import logging
import tempfile
import json

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app, origins=['*'], methods=['GET', 'POST', 'OPTIONS'], allow_headers=['Content-Type', 'Authorization'])

# 配置
TEMPLATE_PATH = 'api/api/template.xlsx'

@app.route('/health', methods=['GET'])
def health_check():
    """健康检查接口"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'message': '合同管理后端服务运行正常',
        'platform': 'Render'
    })

@app.route('/api/generate-contract', methods=['GET', 'POST', 'OPTIONS'])
def generate_contract():
    """生成合同Excel文件"""
    # 处理OPTIONS请求（CORS预检）
    if request.method == 'OPTIONS':
        return '', 200
    
    # 处理GET请求（健康检查）
    if request.method == 'GET':
        return jsonify({
            'status': 'ok',
            'message': 'Contract API is running',
            'timestamp': datetime.now().isoformat(),
            'platform': 'Render'
        })
    
    # 处理POST请求
    try:
        # 获取请求数据
        data = request.get_json()
        logger.info(f"收到合同数据: {data}")
        
        if not data:
            return jsonify({'error': '未收到数据'}), 400
        
        # 验证必要字段
        if not data.get('buyerName') or not data.get('contractNumber'):
            return jsonify({'error': 'Missing required fields: buyerName, contractNumber'}), 400
        
        # 加载模板文件
        if not os.path.exists(TEMPLATE_PATH):
            return jsonify({'error': '模板文件不存在'}), 404
        
        # 使用openpyxl加载工作簿
        workbook = load_workbook(TEMPLATE_PATH)
        
        # 获取两个sheet
        sc_sheet = workbook['SC']  # SC sheet
        pi_sheet = workbook['PI']  # PI sheet
        
        # 填充买方信息
        buyer_name = data.get('buyerName', '')
        buyer_phone = data.get('buyerPhone', '')
        buyer_address = data.get('buyerAddress', '')
        
        # 填充卖方信息
        seller_name = data.get('sellerName', '')
        seller_phone = data.get('sellerPhone', '')
        seller_address = data.get('sellerAddress', '')
        
        # 填充合同信息
        contract_number = data.get('contractNumber', '')
        contract_date_raw = data.get('contractDate', '')
        contract_location = data.get('contractLocation', '')
        
        # 处理日期格式：将 YYYY-MM-DD 转换为 YYYY.MM.DD
        if contract_date_raw:
            try:
                date_obj = datetime.strptime(contract_date_raw, '%Y-%m-%d')
                contract_date = date_obj.strftime('%Y.%m.%d')
            except:
                contract_date = contract_date_raw
        else:
            contract_date = datetime.now().strftime('%Y.%m.%d')
        
        # 填充开户行信息
        bank_info = data.get('bankInfo', '')
        
        # 全新的模板数据填充方法
        def fill_template_data(sheet, data_mappings):
            """
            使用全新的方法填充模板数据，完全避免合并单元格问题
            """
            success_count = 0
            total_count = len(data_mappings)
            
            for mapping in data_mappings:
                try:
                    cell_ref = mapping['cell']
                    value = mapping['value']
                    
                    # 方法1: 直接使用sheet的索引访问
                    try:
                        sheet[cell_ref] = value
                        success_count += 1
                        logger.debug(f"成功设置 {cell_ref}: {value}")
                        continue
                    except Exception as e1:
                        logger.debug(f"方法1失败 {cell_ref}: {e1}")
                    
                    # 方法2: 使用cell()方法
                    try:
                        from openpyxl.utils import column_index_from_string
                        col = column_index_from_string(cell_ref.replace('0', '').replace('1', '').replace('2', '').replace('3', '').replace('4', '').replace('5', '').replace('6', '').replace('7', '').replace('8', '').replace('9', ''))
                        row = int(''.join(filter(str.isdigit, cell_ref)))
                        cell = sheet.cell(row=row, column=col)
                        cell.value = value
                        success_count += 1
                        logger.debug(f"方法2成功设置 {cell_ref}: {value}")
                        continue
                    except Exception as e2:
                        logger.debug(f"方法2失败 {cell_ref}: {e2}")
                    
                    # 方法3: 使用坐标转换
                    try:
                        from openpyxl.utils import coordinate_from_string
                        coord = coordinate_from_string(cell_ref)
                        cell = sheet.cell(row=coord[1], column=coord[0])
                        cell.value = value
                        success_count += 1
                        logger.debug(f"方法3成功设置 {cell_ref}: {value}")
                        continue
                    except Exception as e3:
                        logger.debug(f"方法3失败 {cell_ref}: {e3}")
                    
                    logger.error(f"所有方法都失败: {cell_ref}")
                    
                except Exception as e:
                    logger.error(f"填充数据失败 {mapping.get('cell', 'unknown')}: {e}")
            
            logger.info(f"数据填充完成: {success_count}/{total_count}")
            return success_count
        
        # 创建模板副本的方法
        def create_template_copy():
            """
            创建模板的完整副本，避免修改原始模板
            """
            try:
                # 加载原始模板
                template_workbook = openpyxl.load_workbook(TEMPLATE_PATH)
                
                # 创建新的工作簿
                new_workbook = openpyxl.Workbook()
                
                # 复制所有工作表
                for sheet_name in template_workbook.sheetnames:
                    if sheet_name in new_workbook.sheetnames:
                        new_workbook.remove(new_workbook[sheet_name])
                    
                    # 复制工作表
                    source_sheet = template_workbook[sheet_name]
                    new_sheet = new_workbook.create_sheet(title=sheet_name)
                    
                    # 复制所有单元格数据
                    for row in source_sheet.iter_rows():
                        for cell in row:
                            new_sheet[cell.coordinate] = cell.value
                    
                    # 复制合并单元格
                    for merged_range in source_sheet.merged_cells.ranges:
                        new_sheet.merge_cells(str(merged_range))
                    
                    # 复制行高和列宽
                    for row_num in range(1, source_sheet.max_row + 1):
                        if source_sheet.row_dimensions[row_num].height:
                            new_sheet.row_dimensions[row_num].height = source_sheet.row_dimensions[row_num].height
                    
                    for col_num in range(1, source_sheet.max_column + 1):
                        col_letter = openpyxl.utils.get_column_letter(col_num)
                        if source_sheet.column_dimensions[col_letter].width:
                            new_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width
                
                # 删除默认的Sheet
                if 'Sheet' in new_workbook.sheetnames:
                    new_workbook.remove(new_workbook['Sheet'])
                
                return new_workbook
                
            except Exception as e:
                logger.error(f"创建模板副本失败: {e}")
                return None
        
        # 创建模板副本
        workbook = create_template_copy()
        if not workbook:
            return jsonify({'error': '创建模板副本失败'}), 500
        
        # 获取工作表
        sc_sheet = workbook['SC']
        pi_sheet = workbook['PI']
        
        # 准备基础信息数据
        basic_cell_mappings = [
            {'cell': 'C3', 'value': buyer_name},
            {'cell': 'C4', 'value': buyer_address},
            {'cell': 'C5', 'value': buyer_phone},
            {'cell': 'C6', 'value': seller_name},
            {'cell': 'C7', 'value': seller_address},
            {'cell': 'C8', 'value': seller_phone},
            {'cell': 'G3', 'value': contract_number},
            {'cell': 'G4', 'value': contract_date},
            {'cell': 'G5', 'value': contract_location},
            {'cell': 'E7', 'value': bank_info}
        ]
        
        # 使用新方法填充基础信息
        basic_success = fill_template_data(sc_sheet, basic_cell_mappings)
        logger.info(f"基础信息填充成功: {basic_success}/{len(basic_cell_mappings)}")
        
        # 处理货物信息
        goods_data = data.get('goodsData', [])
        logger.info(f"收到的货物数据: {goods_data}")
        if goods_data:
            # 先处理行的显示/隐藏逻辑
            goods_count = len(goods_data)
            
            # 默认隐藏所有货物行（第11-20行）在两个sheet中
            for row_num in range(11, 21):
                sc_sheet.row_dimensions[row_num].hidden = True
                pi_sheet.row_dimensions[row_num].hidden = True
            
            # 根据实际货物行数显示对应的行在两个sheet中
            for i in range(min(goods_count, 10)):  # 最多支持10行货物
                row_num = 11 + i
                sc_sheet.row_dimensions[row_num].hidden = False
                pi_sheet.row_dimensions[row_num].hidden = False
                
                # 只填充货物数据到SC sheet，PI sheet保持原样
                goods_item = goods_data[i]
                goods_cell_mappings = [
                    {'cell': f'B{row_num}', 'value': goods_item.get('model', goods_item.get('name', ''))},
                    {'cell': f'C{row_num}', 'value': goods_item.get('description', goods_item.get('specification', ''))},
                    {'cell': f'D{row_num}', 'value': goods_item.get('color', '')},  # 颜色字段
                    {'cell': f'E{row_num}', 'value': goods_item.get('quantity', '')},
                    {'cell': f'F{row_num}', 'value': goods_item.get('unitPrice', '')},
                    {'cell': f'G{row_num}', 'value': goods_item.get('totalAmount', goods_item.get('amount', ''))}
                ]
                fill_template_data(sc_sheet, goods_cell_mappings)
        
        # 处理运输信息 - 使用新方法
        transport_cell_mappings = []
        
        # D21-E21 - 装运港（合并单元格）
        port_of_loading = data.get('portOfLoading', '')
        if port_of_loading:
            transport_cell_mappings.append({'cell': 'D21', 'value': port_of_loading})
        
        # D22-E22 - 目的港（合并单元格）
        final_destination = data.get('finalDestination', '')
        if final_destination:
            transport_cell_mappings.append({'cell': 'D22', 'value': final_destination})
        
        # B23-G23 - 金额大写
        amount_in_words = data.get('amountInWords', '')
        if amount_in_words:
            transport_cell_mappings.append({'cell': 'B23', 'value': amount_in_words})
        
        # D24-G24 - 付款条件（合并单元格）
        payment_terms = data.get('paymentTerms', '')
        if payment_terms:
            transport_cell_mappings.append({'cell': 'D24', 'value': payment_terms})
        
        # D25-G25 - 运输路线（合并单元格）
        transport_route = data.get('transportRoute', '')
        if transport_route:
            transport_cell_mappings.append({'cell': 'D25', 'value': transport_route})
        
        # D26-G26 - 运输方式（合并单元格）
        mode_of_shipment = data.get('modeOfShipment', '')
        if mode_of_shipment:
            # 转换运输方式为中文+英文
            shipment_mapping = {
                'SEA': '海运 SEA',
                'LAND': '陆运 LAND',
                'AIR': '空运 AIR'
            }
            shipment_text = shipment_mapping.get(mode_of_shipment.upper(), mode_of_shipment)
            transport_cell_mappings.append({'cell': 'D26', 'value': shipment_text})
        
        # 使用新方法填充运输信息
        if transport_cell_mappings:
            transport_success = fill_template_data(sc_sheet, transport_cell_mappings)
            logger.info(f"运输信息填充成功: {transport_success}/{len(transport_cell_mappings)}")
        
        # 生成输出文件名 - 使用合同编号
        if contract_number:
            # 清理合同编号中的特殊字符，确保文件名安全
            safe_contract_number = "".join(c for c in contract_number if c.isalnum() or c in ('-', '_'))
            output_filename = f'{safe_contract_number}_Contract.xlsx'
        else:
            # 如果没有合同编号，使用时间戳作为备用
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f'contract_{timestamp}.xlsx'
        
        # 创建临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            output_path = tmp_file.name
            workbook.save(output_path)
            logger.info(f"合同文件已生成: {output_path}")
            
            # 返回文件下载
            return send_file(
                output_path,
                as_attachment=True,
                download_name=output_filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
    except Exception as e:
        logger.error(f"生成合同时发生错误: {str(e)}")
        return jsonify({'error': f'生成合同失败: {str(e)}'}), 500

@app.route('/api/template-info', methods=['GET'])
def get_template_info():
    """获取模板文件信息"""
    try:
        if os.path.exists(TEMPLATE_PATH):
            file_size = os.path.getsize(TEMPLATE_PATH)
            return jsonify({
                'status': 'ok',
                'template_exists': True,
                'file_size': file_size,
                'file_path': TEMPLATE_PATH
            })
        else:
            return jsonify({
                'status': 'error',
                'template_exists': False,
                'message': '模板文件不存在'
            }), 404
    except Exception as e:
        return jsonify({'error': f'获取模板信息失败: {str(e)}'}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False) 